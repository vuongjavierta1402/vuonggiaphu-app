#!/usr/bin/env python3
"""
tdm_crawl.py — Crawl all products from tdm.vn and export to Excel
in VGP admin import format.

Usage:
    python tdm_crawl.py                        # crawl / auto-resume all ~457 products
    python tdm_crawl.py --limit 20             # quick test (first 20 only)
    python tdm_crawl.py --output products.xlsx # custom output filename

Progress is saved to tdm_checkpoint.json every 10 products.
The script always auto-resumes from that file if it exists —
no extra flag needed.

Requirements (Python 3.9+):
    pip install requests beautifulsoup4 lxml openpyxl
"""

import argparse
import json
import logging
import random
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── Configuration ─────────────────────────────────────────────────────────────

SITEMAP_URL    = "https://www.tdm.vn/sitemaps/en-gb.sitemap.products.xml"
BASE_URL       = "https://www.tdm.vn"
CHECKPOINT     = Path("tdm_checkpoint.json")
DEFAULT_OUTPUT = Path("tdm_products.xlsx")   # stable name beside the checkpoint

DELAY_MIN      = 1.0   # seconds between normal requests
DELAY_MAX      = 3.0
RESET_WAIT_MIN = 25    # seconds to wait after a 400 / session reset
RESET_WAIT_MAX = 40
SAVE_EVERY     = 10    # save checkpoint + Excel every N products

# Pool of User-Agents to rotate on session reset
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

BASE_HEADERS = {
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,"
                       "image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection":      "keep-alive",
    "Referer":         "https://www.tdm.vn/",
    "Sec-Fetch-Dest":  "document",
    "Sec-Fetch-Mode":  "navigate",
    "Sec-Fetch-Site":  "same-origin",
    "Sec-Fetch-User":  "?1",
    "DNT":             "1",
}

# ── Category mapping: TDM breadcrumb keywords → VGP (category, subcategory) ──
# Rules: longer / more-specific terms first; all keywords lowercase.

CATEGORY_MAP = [
    # ── Thiết Bị Vệ Sinh ──────────────────────────────────────────────────
    ("bồn cầu thông minh",   "Thiết Bị Vệ Sinh", "Bồn cầu điện tử"),
    ("bồn cầu điện tử",      "Thiết Bị Vệ Sinh", "Bồn cầu điện tử"),
    ("bồn cầu tự động",      "Thiết Bị Vệ Sinh", "Bồn cầu điện tử"),
    ("bàn cầu cảm ứng",      "Thiết Bị Vệ Sinh", "Bồn cầu điện tử"),
    ("nắp bồn cầu",          "Thiết Bị Vệ Sinh", "Nắp bồn cầu"),
    ("nắp bàn cầu",          "Thiết Bị Vệ Sinh", "Nắp bồn cầu"),
    ("bồn cầu treo tường",   "Thiết Bị Vệ Sinh", "Bồn cầu"),
    ("bồn cầu",              "Thiết Bị Vệ Sinh", "Bồn cầu"),
    ("bàn cầu",              "Thiết Bị Vệ Sinh", "Bồn cầu"),
    ("chậu rửa lavabo",      "Thiết Bị Vệ Sinh", "Chậu Lavabo"),
    ("chậu rửa mặt",         "Thiết Bị Vệ Sinh", "Chậu Lavabo"),
    ("chậu lavabo",          "Thiết Bị Vệ Sinh", "Chậu Lavabo"),
    ("lavabo cổ điển",       "Thiết Bị Vệ Sinh", "Chậu Lavabo"),
    ("lavabo",               "Thiết Bị Vệ Sinh", "Chậu Lavabo"),
    ("vòi lavabo",           "Thiết Bị Vệ Sinh", "Vòi chậu"),
    ("vòi chậu",             "Thiết Bị Vệ Sinh", "Vòi chậu"),
    ("vòi cảm ứng",          "Thiết Bị Vệ Sinh", "Vòi chậu"),
    ("vòi sen cây",          "Thiết Bị Vệ Sinh", "Sen cây"),
    ("sen cây",              "Thiết Bị Vệ Sinh", "Sen cây"),
    ("cây sen",              "Thiết Bị Vệ Sinh", "Sen cây"),
    ("sen âm tường",         "Thiết Bị Vệ Sinh", "Sen cây"),
    ("vòi sen tắm",          "Thiết Bị Vệ Sinh", "Vòi sen"),
    ("vòi sen",              "Thiết Bị Vệ Sinh", "Vòi sen"),
    ("sen tắm",              "Thiết Bị Vệ Sinh", "Vòi sen"),
    ("củ sen",               "Thiết Bị Vệ Sinh", "Vòi sen"),
    ("vòi bồn tắm",          "Thiết Bị Vệ Sinh", "Bồn tắm"),
    ("bồn tắm",              "Thiết Bị Vệ Sinh", "Bồn tắm"),
    ("van xả tiểu",          "Thiết Bị Vệ Sinh", "Bồn tiểu"),
    ("bồn tiểu nam",         "Thiết Bị Vệ Sinh", "Bồn tiểu"),
    ("bồn tiểu nữ",          "Thiết Bị Vệ Sinh", "Bồn tiểu"),
    ("bồn tiểu",             "Thiết Bị Vệ Sinh", "Bồn tiểu"),
    ("phòng tắm kính",       "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("gương phòng tắm",      "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("gương trang điểm",     "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("thanh vịn",            "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phễu thoát sàn",       "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phểu thoát sàn",       "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("vòi xịt",              "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phụ kiện phòng tắm",   "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("bộ phụ kiện",          "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phụ kiện cotto",       "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phụ kiện caesar",      "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phụ kiện inax",        "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phụ kiện toto",        "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("phụ kiện",             "Thiết Bị Vệ Sinh", "Phụ kiện"),
    ("lắp đặt thiết bị vệ sinh", "Thiết Bị Vệ Sinh", ""),
    ("thiết bị vệ sinh",     "Thiết Bị Vệ Sinh", ""),

    # ── Thiết Bị Nhà Bếp ─────────────────────────────────────────────────
    ("bếp điện từ",          "Thiết Bị Nhà Bếp", "Bếp Điện Từ"),
    ("bếp từ",               "Thiết Bị Nhà Bếp", "Bếp Điện Từ"),
    ("bếp điện",             "Thiết Bị Nhà Bếp", "Bếp Điện Từ"),
    ("bếp gas",              "Thiết Bị Nhà Bếp", "Bếp Gas"),
    ("máy hút mùi",          "Thiết Bị Nhà Bếp", "Máy Hút Mùi"),
    ("máy hút khói",         "Thiết Bị Nhà Bếp", "Máy Hút Mùi"),
    ("lò nướng",             "Thiết Bị Nhà Bếp", "Lò nướng"),
    ("lò vi sóng",           "Thiết Bị Nhà Bếp", "Lò vi sóng"),
    ("máy rửa chén",         "Thiết Bị Nhà Bếp", "Máy rửa chén"),
    ("tủ lạnh",              "Thiết Bị Nhà Bếp", "Tủ lạnh"),
    ("bồn rửa chén",         "Thiết Bị Nhà Bếp", "Chậu rửa chén"),
    ("chậu rửa chén",        "Thiết Bị Nhà Bếp", "Chậu rửa chén"),
    ("chậu rửa bát",         "Thiết Bị Nhà Bếp", "Chậu rửa chén"),
    ("chậu rửa inox",        "Thiết Bị Nhà Bếp", "Chậu rửa chén"),
    ("vòi rửa chén",         "Thiết Bị Nhà Bếp", "Vòi rửa chén"),
    ("vòi bếp",              "Thiết Bị Nhà Bếp", "Vòi rửa chén"),
    ("thiết bị bếp",         "Thiết Bị Nhà Bếp", ""),
    ("thiết bị nhà bếp",     "Thiết Bị Nhà Bếp", ""),

    # ── Thiết Bị Nước ─────────────────────────────────────────────────────
    ("lõi lọc nước",         "Thiết Bị Nước",    "Máy lọc nước"),
    ("lõi máy lọc",          "Thiết Bị Nước",    "Máy lọc nước"),
    ("máy lọc nước",         "Thiết Bị Nước",    "Máy lọc nước"),
    ("thái dương năng",      "Thiết Bị Nước",    "Máy nước nóng"),
    ("nước nóng năng lượng", "Thiết Bị Nước",    "Máy nước nóng"),
    ("năng lượng mặt trời",  "Thiết Bị Nước",    "Máy nước nóng"),
    ("máy nước nóng",        "Thiết Bị Nước",    "Máy nước nóng"),
    ("máy bơm nước",         "Thiết Bị Nước",    "Máy bơm nước"),
    ("bơm nước",             "Thiết Bị Nước",    "Máy bơm nước"),
    ("bồn inox",             "Thiết Bị Nước",    "Bồn nước"),
    ("bồn nhựa",             "Thiết Bị Nước",    "Bồn nước"),
    ("bồn công nghiệp",      "Thiết Bị Nước",    "Bồn nước"),
    ("bồn tự hoại",          "Thiết Bị Nước",    "Bồn tự hoại"),
    ("bồn nước",             "Thiết Bị Nước",    "Bồn nước"),
    ("vật liệu nước",        "Thiết Bị Nước",    ""),
    ("thiết bị nước",        "Thiết Bị Nước",    ""),

    # ── Thiết Bị Điện ─────────────────────────────────────────────────────
    ("bóng đèn led",         "Thiết Bị Điện",    "Đèn Led"),
    ("đèn led âm trần",      "Thiết Bị Điện",    "Đèn Led"),
    ("đèn led",              "Thiết Bị Điện",    "Đèn Led"),
    ("công tắc ổ cắm",       "Thiết Bị Điện",    "Công tắc ổ cắm"),
    ("ổ cắm điện",           "Thiết Bị Điện",    "Công tắc ổ cắm"),
    ("công tắc",             "Thiết Bị Điện",    "Công tắc ổ cắm"),

    # ── Khóa Cửa ──────────────────────────────────────────────────────────
    ("khóa điện tử",         "Khóa Cửa",         "Khóa điện tử"),
    ("khóa cửa chính",       "Khóa Cửa",         "Khóa cửa"),
    ("khóa cửa phòng",       "Khóa Cửa",         "Khóa cửa"),
    ("khóa cửa",             "Khóa Cửa",         "Khóa cửa"),
    ("két sắt",              "Khóa Cửa",         "Két sắt"),
]

BRAND_NORMALIZE = {
    "inax":              "INAX",
    "toto":              "TOTO",
    "caesar":            "Caesar",
    "viglacera":         "VIGLACERA",
    "american standard": "AMERICAN STANDARD",
    "attax":             "ATTAX",
    "haphako":           "HAPHAKO",
    "grohe":             "GROHE",
    "ferroli":           "FERROLI",
    "ariston":           "ARISTON",
    "bosch":             "BOSCH",
}

# ── Logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("tdm_crawl.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("tdm")

# ── Session management ────────────────────────────────────────────────────────
# Kept as a mutable container so _reset_session() can replace it anywhere.

_session_holder = {"session": None}


def _make_session():
    s = requests.Session()
    s.headers.update(BASE_HEADERS)
    s.headers["User-Agent"] = random.choice(USER_AGENTS)
    return s


def _reset_session(reason=""):
    """Drop all cookies and rotate User-Agent — same effect as restarting."""
    _session_holder["session"] = _make_session()
    ua_short = _session_holder["session"].headers["User-Agent"][:60]
    log.info(f"  [session reset{f' ({reason})' if reason else ''}  UA: {ua_short}]")


def _session():
    return _session_holder["session"]


# ── HTTP fetch ────────────────────────────────────────────────────────────────

def _get(url, retries=4):
    """
    Fetch URL → BeautifulSoup, or None after all retries.

    On HTTP 400 the server has fingerprinted our session.  We immediately
    reset (new session + new User-Agent) and wait ~30 s before retrying —
    identical to what a manual restart achieves.
    """
    for attempt in range(retries):
        try:
            r = _session().get(url, timeout=20)

            if r.status_code == 400:
                wait = random.uniform(RESET_WAIT_MIN, RESET_WAIT_MAX)
                log.warning(
                    f"  400 on attempt {attempt+1} — session blocked. "
                    f"Resetting and waiting {wait:.0f}s …"
                )
                _reset_session("400")
                time.sleep(wait)
                continue  # retry with fresh session

            r.raise_for_status()
            r.encoding = r.apparent_encoding or "utf-8"
            return BeautifulSoup(r.text, "lxml")

        except requests.HTTPError:
            raise  # already handled 400 above; other HTTP errors bubble up
        except Exception as exc:
            wait = 8 * (attempt + 1)
            log.warning(f"  attempt {attempt+1}/{retries}: {exc}  (wait {wait}s)")
            time.sleep(wait)

    log.error(f"  FAILED after {retries} attempts: {url}")
    return None


# ── Sitemap ───────────────────────────────────────────────────────────────────

def get_product_urls():
    log.info("Fetching product sitemap …")
    try:
        r = _session().get(SITEMAP_URL, timeout=20)
        r.raise_for_status()
        root = ET.fromstring(r.content)
        ns   = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
        urls = [loc.text.strip() for loc in root.findall(".//sm:loc", ns) if loc.text]
        log.info(f"Sitemap: {len(urls)} product URLs found")
        return urls
    except Exception as exc:
        log.error(f"Sitemap fetch failed: {exc}")
        return []


# ── Helpers ───────────────────────────────────────────────────────────────────

def clean_price(text):
    """'3.650.000 đ' → 3650000"""
    digits = re.sub(r"[^\d]", "", text or "")
    return int(digits) if digits else 0


def map_category(crumbs):
    combined = " | ".join(c.lower() for c in crumbs)
    for keyword, cat, subcat in CATEGORY_MAP:
        if keyword in combined:
            return cat, subcat

    # Fallback: use the raw breadcrumb text rather than returning empty strings.
    # Filter out very long crumbs (usually product names that slipped through).
    short = [c.strip() for c in crumbs if c.strip() and len(c.strip()) < 60]
    if len(short) >= 2:
        return short[0], short[1]
    if len(short) == 1:
        return short[0], ""
    if crumbs:
        return crumbs[0].strip(), ""
    return "", ""


def normalize_brand(raw):
    return BRAND_NORMALIZE.get(raw.strip().lower(), raw.strip())


# ── Product parser ────────────────────────────────────────────────────────────

def parse_product(url):
    """Scrape one product detail page. Returns dict or None."""
    soup = _get(url)
    if not soup:
        return None

    # ── Name ──────────────────────────────────────────────────────────────────
    h1   = soup.select_one("h1")
    name = h1.get_text(strip=True) if h1 else ""
    if not name:
        log.warning(f"  No name at {url}")
        return None

    # ── Prices ────────────────────────────────────────────────────────────────
    # <tr class="old-price"><td>...<span>140.140.000 đ</span></td></tr>
    # <tr class="new-price"><td><h3>96.050.000 đ</h3></td></tr>
    price, discount_price, is_sale = 0, 0, False
    old_row = soup.select_one("tr.old-price")
    new_row = soup.select_one("tr.new-price")
    old_el  = old_row.select_one("span") if old_row else None
    new_el  = new_row.select_one("h3")   if new_row else None
    if old_el and new_el:
        price          = clean_price(old_el.get_text())
        discount_price = clean_price(new_el.get_text())
        is_sale        = True
    elif new_el:
        price = clean_price(new_el.get_text())
    elif old_el:
        price = clean_price(old_el.get_text())

    if price == 0:
        for el in soup.select("[class*='price'], .gia, .product-price"):
            val = clean_price(el.get_text())
            if val > 0:
                price = val
                break

    # ── Images (returned as a list) ───────────────────────────────────────────
    # Real URLs are in data-flickity-lazyload on img.carousel-cell-image.
    # Two sizes exist: 1090x1090 (full) and 74x74 (thumbnail) — keep full only.
    seen, images = set(), []

    def _add(u):
        if u and u not in seen:
            seen.add(u); images.append(u)

    for img in soup.select("img.carousel-cell-image"):
        raw = (img.get("data-flickity-lazyload") or "").strip()
        if raw and "1090x1090" in raw:
            _add(raw if raw.startswith("http") else urljoin(BASE_URL, raw))

    # Fallback: accept any size if no 1090x1090 images found
    if not images:
        for img in soup.select("img.carousel-cell-image"):
            raw = (img.get("data-flickity-lazyload") or "").strip()
            if raw:
                _add(raw if raw.startswith("http") else urljoin(BASE_URL, raw))

    images = images[:8]   # list of URL strings

    # ── Brand ─────────────────────────────────────────────────────────────────
    # The manufacturer logo is inside an element with class "manufacturers_img".
    # Use alt first, fall back to title.
    brand = ""
    mfr_img = soup.select_one(".manufacturers_img img, .manufacturer_img img")
    if mfr_img:
        raw = mfr_img.get("alt", "").strip() or mfr_img.get("title", "").strip()
        brand = normalize_brand(raw)

    # Fallback: any img whose src path contains a known brand keyword
    if not brand:
        for img_el in soup.find_all("img", src=True):
            src_lower = img_el["src"].lower()
            for key in BRAND_NORMALIZE:
                if key.replace(" ", "-") in src_lower or key.replace(" ", "_") in src_lower:
                    brand = BRAND_NORMALIZE[key]
                    break
            if brand:
                break

    # ── SKU / productCode ─────────────────────────────────────────────────────
    product_code = ""

    # Strategy 1: "Mã sản phẩm:" label → adjacent strong/span
    for node in soup.find_all(string=re.compile(r"Mã\s*s[aả]n\s*ph[aả]m|Mã\s*SP|SKU", re.I)):
        parent = node.parent
        tail   = re.sub(
            r"(?i).*?(Mã\s*s[aả]n\s*ph[aả]m|Mã\s*SP|SKU)\s*[:\-]?\s*", "", str(node)
        ).strip()
        if not tail:
            nxt  = parent.find_next(["strong", "span", "td", "b"])
            tail = nxt.get_text(strip=True) if nxt else ""
        code = re.sub(r"\s+", "", tail)
        if code and 2 <= len(code) <= 40:
            product_code = code.replace("/", "-")
            break

    # Strategy 2: regex over li/td/span text
    if not product_code:
        for el in soup.select("li, td, p, span"):
            txt = el.get_text(separator=" ", strip=True)
            m   = re.search(r"(?:Mã|SKU)\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-\/\.]{2,30})", txt, re.I)
            if m:
                product_code = m.group(1).strip().replace("/", "-")
                break

    # Fallback: URL slug
    if not product_code:
        slug = url.rstrip("/")
        if slug.endswith(".html"):
            slug = slug[:-5]
        product_code = slug.split("/")[-1].upper()[:40]

    # ── Breadcrumb → category / subcategory ───────────────────────────────────
    crumbs = [
        li.get_text(strip=True)
        for li in soup.select("ul.breadcrumb li")
        if li.get_text(strip=True).lower() not in ("trang chủ", "home")
    ]
    if crumbs and crumbs[-1].lower() == name.lower():
        crumbs = crumbs[:-1]

    category, subcategory = map_category(crumbs)

    # ── Description ───────────────────────────────────────────────────────────
    desc_el = soup.select_one(
        "div.description, #tab-description, #description, "
        ".product-description, .desc-content, .tab-content"
    )
    description = str(desc_el) if desc_el else ""

    return {
        "productCode":   product_code,
        "name":          name,
        "price":         price,
        "discountPrice": discount_price,
        "brand":         brand,
        "category":      category,
        "subcategory":   subcategory,
        "quantity":      0,
        "description":   description,
        "images":        images,
        "sale":          "TRUE" if is_sale else "FALSE",
        "highlighted":   "FALSE",
        "isDisplay":     "TRUE",
    }


# ── Excel export ──────────────────────────────────────────────────────────────

COLUMNS = [
    "productCode", "name", "price", "discountPrice", "brand",
    "category", "subcategory", "quantity", "description",
    "images", "sale", "highlighted", "isDisplay",
]
COL_WIDTHS = {
    "productCode": 22, "name": 55, "price": 16, "discountPrice": 16,
    "brand": 18, "category": 28, "subcategory": 28, "quantity": 10,
    "description": 70, "images": 90, "sale": 10, "highlighted": 13, "isDisplay": 13,
}


def save_excel(products, filepath):
    wb    = Workbook()
    ws    = wb.active
    ws.title = "Products"

    hfill  = PatternFill("solid", fgColor="1E293B")
    hfont  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for ci, col in enumerate(COLUMNS, 1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = halign
        ws.column_dimensions[cell.column_letter].width = COL_WIDTHS.get(col, 18)
    ws.row_dimensions[1].height = 22

    for ri, p in enumerate(products, 2):
        for ci, col in enumerate(COLUMNS, 1):
            val = p.get(col, "")
            if isinstance(val, list):
                val = ",".join(val)
            ws.cell(row=ri, column=ci, value=val)

    ws.freeze_panes = "A2"
    wb.save(filepath)


def _flush(done, output):
    """Save checkpoint JSON + Excel atomically."""
    products = list(done.values())

    # Write checkpoint
    tmp_cp = CHECKPOINT.with_suffix(".tmp")
    tmp_cp.write_text(json.dumps(done, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_cp.replace(CHECKPOINT)

    # Write Excel
    tmp_xl = Path(str(output) + ".tmp")
    save_excel(products, tmp_xl)
    tmp_xl.replace(output)

    log.info(f"  [saved {len(products)} products → {output}  |  checkpoint → {CHECKPOINT}]")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Crawl tdm.vn → Excel (VGP format)")
    parser.add_argument("--limit",  type=int, default=None,
                        help="Max products to crawl (default: all)")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT),
                        help=f"Output Excel path (default: {DEFAULT_OUTPUT})")
    args   = parser.parse_args()
    output = Path(args.output)

    # Initialise session
    _reset_session("startup")

    log.info("=" * 60)
    log.info("  TDM.VN Product Crawler")
    log.info(f"  Output : {output}")
    log.info(f"  Checkpoint : {CHECKPOINT}")
    log.info("=" * 60)

    # 1. Sitemap
    all_urls = get_product_urls()
    if not all_urls:
        log.error("No product URLs found. Aborting.")
        sys.exit(1)

    # 2. Always auto-resume from checkpoint if it exists
    done = {}
    if CHECKPOINT.exists():
        try:
            done = json.loads(CHECKPOINT.read_text(encoding="utf-8"))
            log.info(f"Auto-resumed: {len(done)} products already in checkpoint")
        except Exception as exc:
            log.warning(f"Checkpoint unreadable ({exc}) — starting fresh")

    # 3. Pending URLs
    pending = [u for u in all_urls if u not in done]
    if args.limit:
        pending = pending[:args.limit]

    log.info(f"Total in sitemap : {len(all_urls)}")
    log.info(f"Already done     : {len(done)}")
    log.info(f"To crawl now     : {len(pending)}")
    log.info("")

    if not pending:
        log.info("Nothing left to crawl.")
        _flush(done, output)
        return

    # 4. Crawl loop
    failed = []
    for i, url in enumerate(pending, 1):
        log.info(f"[{i:>4}/{len(pending)}]  {url}")
        product = parse_product(url)

        if product:
            done[url] = product
            log.info(
                f"         ✓  {product['productCode']:<22} "
                f"{product['price']:>12,}đ  {product['name'][:45]}"
            )
        else:
            failed.append(url)
            log.warning("         ✗  skipped")

        # Flush every SAVE_EVERY products
        if i % SAVE_EVERY == 0:
            _flush(done, output)

        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    # 5. Final flush
    _flush(done, output)

    # 6. Summary
    print()
    print("=" * 60)
    print(f"  Output      : {output}  ({len(done)} products)")
    print(f"  Checkpoint  : {CHECKPOINT}")
    print(f"  Failed      : {len(failed)}")
    if failed:
        print()
        print("  Failed URLs (re-run to retry automatically):")
        for u in failed:
            print(f"    {u}")
    print("=" * 60)


if __name__ == "__main__":
    main()
